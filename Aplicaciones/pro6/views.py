from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.http import HttpResponse
import csv
from openpyxl import Workbook
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from io import BytesIO
from .models import Encargado, Estructura, PlanMantenimiento, Trabajo

# Create your views here.
def home(request):
    return render(request,"home.html")

# Vista para listar encargados
def listadoEncargados(request):
    encargados = Encargado.objects.all()
    return render(request, 'encargo/listadoEncargado.html', {'encargados': encargados})

# Vista para crear un nuevo encargado
def nuevoEncargado(request):
    return render(request, 'encargo/nuevaEncargado.html')

# Vista para guardar un nuevo encargado
def guardarEncargado(request):
    if request.method == 'POST':
        nombre = request.POST.get('nombre')
        correo = request.POST.get('correo')
        telefono = request.POST.get('telefono')
        puesto = request.POST.get('puesto')
        foto = request.FILES.get('foto')
        
        encargado = Encargado.objects.create(
            nombre=nombre,
            correo=correo,
            telefono=telefono,
            puesto=puesto,
            foto=foto
        )
        messages.success(request, "Encargado guardado exitosamente")
        return redirect('listado_encargados')
    return redirect('nuevoEncargado')

# Vista para editar un encargado
def editarEncargado(request, id):
    encargado = get_object_or_404(Encargado, id=id)
    return render(request, 'encargo/editarEncargado.html', {'encargadoEditar': encargado})

# Vista para actualizar un encargado
def procesoActualizarEncargado(request):
    if request.method == 'POST':
        encargado_id = request.POST.get('id')
        encargado = get_object_or_404(Encargado, id=encargado_id)

        # Actualizar el nombre
        encargado.nombre = request.POST.get('nombre')

        # Verificar si el campo correo está vacío
        correo = request.POST.get('correo')
        if not correo:
            messages.error(request, "El correo es obligatorio")
            return redirect('editarEncargado', id=encargado_id)

        encargado.correo = correo  # Actualizar el correo
        encargado.telefono = request.POST.get('telefono')  # Actualizar el teléfono

        # Verificar si el campo puesto está vacío, asignar un valor por defecto
        puesto = request.POST.get('puesto')
        if not puesto:
            puesto = "Puesto no especificado"  # Valor predeterminado si el campo está vacío
        encargado.puesto = puesto  # Actualizar el puesto

        # Verificar si se subió una nueva fotografía
        if 'fotografia' in request.FILES:
            encargado.foto = request.FILES.get('fotografia')  # Actualizar la foto

        encargado.save()  # Guardar los cambios en la base de datos
        messages.success(request, "Encargado actualizado exitosamente")
        
        # Redirigir al listado de encargados después de actualizar
        return redirect('listado_encargados')
    return redirect('listado_encargados')
# Vista para eliminar un encargado
def eliminarEncargado(request, id):
    encargado = get_object_or_404(Encargado, id=id)
    encargado.delete()
    messages.success(request, "Encargado eliminado exitosamente")
    return redirect('listado_encargados')


# Listar estructuras
def listadoEstructuras(request):
    estructuras = Estructura.objects.all()
    return render(request, 'estructura/listadoEstructura.html', {'estructuras': estructuras})

# Crear nueva estructura
def nuevaEstructura(request):
    return render(request, 'estructura/nuevaEstructura.html')

# Guardar nueva estructura
def guardarEstructura(request):
    if request.method == 'POST':
        nombre = request.POST.get('nombre')
        descripcion = request.POST.get('descripcion')
        ubicacion = request.POST.get('ubicacion')
        fecha_instalacion = request.POST.get('fecha_instalacion')
        foto = request.FILES.get('foto')

        # Crear una nueva instancia de Estructura
        estructura = Estructura(
            nombre=nombre,
            descripcion=descripcion,
            ubicacion=ubicacion,
            fecha_instalacion=fecha_instalacion,
            foto=foto
        )
        estructura.save()
        messages.success(request, "Estructura guardada con éxito")
        return redirect('listado_estructuras')
    return redirect('nuevaEstructura')

# Editar estructura
def editarEstructura(request, id):
    estructuraEditar = get_object_or_404(Estructura, id=id)
    return render(request, 'estructura/editarEstructura.html', {'estructuraEditar': estructuraEditar})

# Actualizar estructura
def procesoActualizarEstructura(request):
    if request.method == 'POST':
        id = request.POST.get('id')
        estructura = get_object_or_404(Estructura, id=id)
        
        estructura.nombre = request.POST.get('nombre')
        estructura.descripcion = request.POST.get('descripcion')
        estructura.ubicacion = request.POST.get('ubicacion')
        estructura.fecha_instalacion = request.POST.get('fecha_instalacion')
        if 'foto' in request.FILES:
            estructura.foto = request.FILES['foto']
        
        estructura.save()
        messages.success(request, "Estructura actualizada con éxito")
        return redirect('listado_estructuras')
    return redirect('listado_estructuras')

# Eliminar estructura
def eliminarEstructura(request, id):
    estructura = get_object_or_404(Estructura, id=id)
    estructura.delete()
    messages.success(request, "Estructura eliminada con éxito")
    return redirect('listado_estructuras')

# Exportar a Excel
def exportarEstructurasExcel(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="estructuras.xlsx"'
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Estructuras"
    
    # Escribir encabezados
    ws.append(['ID', 'Nombre', 'Descripción', 'Ubicación', 'Fecha de Instalación', 'Fotografía'])
    
    estructuras = Estructura.objects.all()
    for estructura in estructuras:
        ws.append([estructura.id, estructura.nombre, estructura.descripcion, estructura.ubicacion, estructura.fecha_instalacion, estructura.foto.url if estructura.foto else ''])
    
    wb.save(response)
    return response

# Exportar a CSV
def exportarEstructurasCSV(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="estructuras.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['ID', 'Nombre', 'Descripción', 'Ubicación', 'Fecha de Instalación', 'Fotografía'])
    
    estructuras = Estructura.objects.all()
    for estructura in estructuras:
        writer.writerow([estructura.id, estructura.nombre, estructura.descripcion, estructura.ubicacion, estructura.fecha_instalacion, estructura.foto.url if estructura.foto else ''])
    
    return response

# Exportar a PDF
def exportarEstructurasPDF(request):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="estructuras.pdf"'
    
    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter
    
    p.setFont("Helvetica", 12)
    y_position = height - 40
    p.drawString(30, y_position, "ID  | Nombre | Descripción | Ubicación | Fecha de Instalación")
    
    y_position -= 20
    estructuras = Estructura.objects.all()
    for estructura in estructuras:
        p.drawString(30, y_position, f"{estructura.id} | {estructura.nombre} | {estructura.descripcion} | {estructura.ubicacion} | {estructura.fecha_instalacion}")
        y_position -= 20
    
    p.showPage()
    p.save()
    
    buffer.seek(0)
    response.write(buffer.read())
    buffer.close()
    
    return response

# Vista para mostrar el listado de trabajos
def listadoTrabajos(request):
    trabajos = Trabajo.objects.all()
    return render(request, 'trabajo/listadoTrabajo.html', {'trabajos': trabajos})

# Vista para crear un nuevo trabajo
def nuevoTrabajo(request):
    estructuras = Estructura.objects.all()  # Obtener todas las estructuras
    return render(request, 'trabajo/nuevaTrabajo.html', {'estructuras': estructuras})

# Vista para guardar un nuevo trabajo
def guardarTrabajo(request):
    if request.method == "POST":
        estructura_id = request.POST.get('estructura')
        descripcion = request.POST.get('descripcion')
        fecha_programada = request.POST.get('fecha_programada')
        fecha_ejecucion = request.POST.get('fecha_ejecucion')
        estado = request.POST.get('estado')
        documento = request.FILES.get('documentos')
        
        estructura = Estructura.objects.get(id=estructura_id)
        
        trabajo = Trabajo(
            estructura=estructura,
            descripcion=descripcion,
            fecha_programada=fecha_programada,
            fecha_ejecucion=fecha_ejecucion if fecha_ejecucion else None,
            estado=estado,
            documentos=documento if documento else None
        )
        trabajo.save()
        messages.success(request, "Trabajo guardado exitosamente")
        return redirect('listado_trabajos')

    return redirect('nuevoTrabajo')

# Vista para editar un trabajo
def editarTrabajo(request, id):
    trabajo = get_object_or_404(Trabajo, id=id)
    estructuras = Estructura.objects.all()
    return render(request, 'trabajo/editarTrabajo.html', {'trabajo': trabajo, 'estructuras': estructuras})

# Vista para procesar la actualización de un trabajo
def procesoActualizarTrabajo(request):
    if request.method == "POST":
        trabajo_id = request.POST.get('id')
        trabajo = get_object_or_404(Trabajo, id=trabajo_id)
        
        estructura_id = request.POST.get('estructura')
        descripcion = request.POST.get('descripcion')
        fecha_programada = request.POST.get('fecha_programada')
        fecha_ejecucion = request.POST.get('fecha_ejecucion')
        estado = request.POST.get('estado')
        documento = request.FILES.get('documento')
        
        estructura = Estructura.objects.get(id=estructura_id)
        
        trabajo.estructura = estructura
        trabajo.descripcion = descripcion
        trabajo.fecha_programada = fecha_programada
        trabajo.fecha_ejecucion = fecha_ejecucion if fecha_ejecucion else None
        trabajo.estado = estado
        
        # Si hay un nuevo documento, se reemplaza el antiguo
        if documento:
            trabajo.documento = documento  # Reemplaza el documento actual

        trabajo.save()
        
        messages.success(request, "Trabajo actualizado exitosamente")
        return redirect('listado_trabajos')
    
    return redirect('listado_trabajos')

# Vista para eliminar un trabajo
def eliminarTrabajo(request, id):
    trabajo = get_object_or_404(Trabajo, id=id)
    trabajo.delete()
    messages.success(request, "Trabajo eliminado exitosamente")
    return redirect('listado_trabajos')

# Exportar a Excel
def exportarTrabajosExcel(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="trabajos.xlsx"'
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Trabajos"
    
    # Escribir encabezados
    ws.append(['ID', 'Estructura', 'Descripción', 'Fecha Programada', 'Fecha Ejecución', 'Estado', 'Documentos'])
    
    trabajos = Trabajo.objects.all()
    for trabajo in trabajos:
        ws.append([trabajo.id, trabajo.estructura.nombre, trabajo.descripcion, trabajo.fecha_programada, trabajo.fecha_ejecucion, trabajo.estado, trabajo.documentos.url if trabajo.documentos else ''])
    
    wb.save(response)
    return response

# Exportar a CSV
def exportarTrabajosCSV(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="trabajos.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['ID', 'Estructura', 'Descripción', 'Fecha Programada', 'Fecha Ejecución', 'Estado', 'Documentos'])
    
    trabajos = Trabajo.objects.all()
    for trabajo in trabajos:
        writer.writerow([trabajo.id, trabajo.estructura.nombre, trabajo.descripcion, trabajo.fecha_programada, trabajo.fecha_ejecucion, trabajo.estado, trabajo.documentos.url if trabajo.documentos else ''])
    
    return response

# Exportar a PDF
def exportarTrabajosPDF(request):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="trabajos.pdf"'
    
    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter
    
    p.setFont("Helvetica", 12)
    y_position = height - 40
    p.drawString(30, y_position, "ID  | Estructura | Descripción | Fecha Programada | Fecha Ejecución | Estado | Documentos")
    
    y_position -= 20
    trabajos = Trabajo.objects.all()
    for trabajo in trabajos:
        p.drawString(30, y_position, f"{trabajo.id} | {trabajo.estructura.nombre} | {trabajo.descripcion} | {trabajo.fecha_programada} | {trabajo.fecha_ejecucion if trabajo.fecha_ejecucion else 'N/A'} | {trabajo.estado} | {trabajo.documentos.url if trabajo.documentos else 'N/A'}")
        y_position -= 20
    
    p.showPage()
    p.save()
    
    buffer.seek(0)
    response.write(buffer.read())
    buffer.close()
    
    return response


from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from .models import PlanMantenimiento, Estructura, Trabajo, Encargado
from django.contrib import messages
import csv
from io import StringIO
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from django.template.loader import render_to_string


# Vista para mostrar el listado de los planes de mantenimiento
def listado_planes_mantenimiento(request):
    planes_mantenimiento = PlanMantenimiento.objects.all()
    return render(request, 'Plan/listadoPlanMantenimiento.html', {'planes_mantenimiento': planes_mantenimiento})

# Vista para crear un nuevo plan de mantenimiento
def nuevaPlanMantenimiento(request):
    estructuras = Estructura.objects.all()
    trabajos = Trabajo.objects.all()
    encargados = Encargado.objects.all()

    return render(request, 'Plan/nuevaPlanMantenimiento.html', {
        'estructuras': estructuras,
        'trabajos': trabajos,
        'encargados': encargados
    })


# Vista para guardar un nuevo plan de mantenimiento
def guardarPlanMantenimiento(request):
    if request.method == 'POST':
        estructura_id = request.POST.get('estructura')
        trabajo_id = request.POST.get('trabajo')
        encargado_id = request.POST.get('encargado')
        fecha_inicio = request.POST.get('fecha_inicio')
        fecha_fin = request.POST.get('fecha_fin')
        tipo = request.POST.get('tipo')
        documento = request.FILES.get('documentos')

        estructura = Estructura.objects.get(id=estructura_id)
        trabajo = Trabajo.objects.get(id=trabajo_id)
        encargado = Encargado.objects.get(id=encargado_id)

        # Crear nuevo plan de mantenimiento
        plan = PlanMantenimiento(
            estructura=estructura,
            trabajo=trabajo,
            encargado=encargado,
            fecha_inicio=fecha_inicio,
            fecha_fin=fecha_fin,
            tipo=tipo,
            documentos=documento
        )
        plan.save()

        messages.success(request, "Plan de mantenimiento guardado con éxito.")
        return redirect('listado_planes_mantenimiento')
    else:
        return redirect('nueva_plan_mantenimiento')

def editarPlanMantenimiento(request, id):
    plan = get_object_or_404(PlanMantenimiento, id=id)
    estructuras = Estructura.objects.all()
    trabajos = Trabajo.objects.all()
    encargados = Encargado.objects.all()

    return render(request, 'Plan/editarPlanMantenimiento.html', {
        'plan': plan,
        'estructuras': estructuras,
        'trabajos': trabajos,
        'encargados': encargados
    })

# Vista para procesar la actualización del plan de mantenimiento
def procesoActualizarPlanMantenimiento(request):
    if request.method == 'POST':
        plan_id = request.POST.get('plan_id')
        plan = get_object_or_404(PlanMantenimiento, id=plan_id)

        estructura_id = request.POST.get('estructura')
        trabajo_id = request.POST.get('trabajo')
        encargado_id = request.POST.get('encargado')
        fecha_inicio = request.POST.get('fecha_inicio')
        fecha_fin = request.POST.get('fecha_fin')
        tipo = request.POST.get('tipo')
        documento = request.FILES.get('documento_plan')

        # Get Estructura, Trabajo, and Encargado objects
        estructura = get_object_or_404(Estructura, id=estructura_id)
        trabajo = get_object_or_404(Trabajo, id=trabajo_id)
        encargado = get_object_or_404(Encargado, id=encargado_id)

        # Update the plan fields
        plan.estructura = estructura
        plan.trabajo = trabajo
        plan.encargado = encargado
        plan.fecha_inicio = fecha_inicio
        plan.fecha_fin = fecha_fin
        plan.tipo = tipo
        
        if documento:
            plan.documentos = documento

        plan.save()
        messages.success(request, "Plan de mantenimiento actualizado con éxito.")
        return redirect('listado_planes_mantenimiento')

    return redirect('listado_planes_mantenimiento')

# Vista para eliminar un plan de mantenimiento
def eliminarPlanMantenimiento(request, id):
    plan = get_object_or_404(PlanMantenimiento, id=id)
    plan.delete()
    messages.success(request, "Plan de mantenimiento eliminado con éxito.")
    return redirect('listado_planes_mantenimiento')


# Exportar a Excel
def exportarPlanesExcel(request):
    planes_mantenimiento = PlanMantenimiento.objects.all()

    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename=planes_mantenimiento.xlsx'

    wb = Workbook()
    ws = wb.active
    ws.title = "Planes de Mantenimiento"
    ws.append(['ID', 'Estructura', 'Trabajo', 'Encargado', 'Fecha Inicio', 'Fecha Fin', 'Tipo', 'Documento'])

    for plan in planes_mantenimiento:
        ws.append([
            plan.id,
            plan.estructura.nombre,
            plan.trabajo.descripcion,
            plan.encargado.nombre,
            plan.fecha_inicio,
            plan.fecha_fin,
            plan.tipo,
            plan.documentos.url if plan.documentos else 'No disponible'
        ])

    wb.save(response)
    return response


# Exportar a CSV
def exportarPlanesCSV(request):
    planes_mantenimiento = PlanMantenimiento.objects.all()

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename=planes_mantenimiento.csv'

    writer = csv.writer(response)
    writer.writerow(['ID', 'Estructura', 'Trabajo', 'Encargado', 'Fecha Inicio', 'Fecha Fin', 'Tipo', 'Documento'])

    for plan in planes_mantenimiento:
        writer.writerow([
            plan.id,
            plan.estructura.nombre,
            plan.trabajo.descripcion,
            plan.encargado.nombre,
            plan.fecha_inicio,
            plan.fecha_fin,
            plan.tipo,
            plan.documentos.url if plan.documentos else 'No disponible'
        ])

    return response

from io import BytesIO
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from django.http import HttpResponse
from .models import PlanMantenimiento  # Asegúrate de importar tu modelo

# Exportar a PDF
def exportarPlanesPDF(request):
    planes_mantenimiento = PlanMantenimiento.objects.all()

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=planes_mantenimiento.pdf'

    # Usar BytesIO para manejar el PDF como binario
    buffer = BytesIO()

    # Crear un objeto canvas de ReportLab
    p = canvas.Canvas(buffer, pagesize=letter)
    text_object = p.beginText(40, 750)
    text_object.setFont("Helvetica", 10)

    # Escribir encabezado
    text_object.textLine('ID | Estructura | Trabajo | Encargado | Fecha Inicio | Fecha Fin | Tipo | Documento')

    # Escribir los datos de los planes de mantenimiento
    for plan in planes_mantenimiento:
        text_object.textLine(f"{plan.id} | {plan.estructura.nombre} | {plan.trabajo.descripcion} | {plan.encargado.nombre} | {plan.fecha_inicio} | {plan.fecha_fin} | {plan.tipo} | {plan.documentos.url if plan.documentos else 'No disponible'}")

    # Dibujar el texto en el PDF
    p.drawText(text_object)
    p.showPage()
    p.save()

    # Obtener el contenido del buffer y escribirlo en la respuesta HTTP
    buffer.seek(0)  # Mover el puntero al inicio del buffer
    response.write(buffer.read())

    return response


from django.core.mail import send_mail
from django.conf import settings
from django.template.loader import render_to_string
from django.utils.html import strip_tags

def enviar_correo_encargado(request, id):
    try:
        # Obtener el encargado por ID
        encargado = get_object_or_404(Encargado, id=id)
        
        # Verificar si el encargado tiene correo electrónico
        if not encargado.correo:
            messages.error(request, 'Este encargado no tiene un correo electrónico asociado.')
            return redirect('listado_encargados')  # Redirigir a la lista de encargados
        
        # Preparar el contenido del correo
        context = {
            'encargado': encargado,
            'BASE_URL': request.build_absolute_uri('/')[:-1]  # Agregar la URL base si es necesario
        }
        
        # Renderizar el template HTML
        html_message = render_to_string('email_template.html', context)
        plain_message = strip_tags(html_message)
        
        # Enviar el correo
        send_mail(
            subject='Detalles del Encargado',
            message=plain_message,
            from_email=settings.DEFAULT_FROM_EMAIL,
            recipient_list=[encargado.correo],
            html_message=html_message,
            fail_silently=False,
        )
        
        messages.success(request, f'Correo enviado exitosamente a {encargado.correo}')
    except Exception as e:
        messages.error(request, f'Error al enviar el correo: {str(e)}')
    
    return redirect('listado_encargados')  # Redirigir a la lista de encargados

